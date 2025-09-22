import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import io
import base64

st.set_page_config(layout="wide")

st.title("Pegado de datos automatizado dentro de un mismo archivo Excel")
st.write("Sube tu archivo para pegar datos de una hoja a otra.")

# --- LÓGICA DE PROCESAMIENTO ---
@st.cache_data
def obtener_nombres_de_hojas(uploaded_file):
    """Función para obtener los nombres de todas las hojas de un archivo de Excel."""
    try:
        xls = pd.ExcelFile(uploaded_file)
        return xls.sheet_names
    except Exception as e:
        st.error(f"No se pudo leer el archivo: {e}")
        return []

def procesar_hoja(single_file, source_sheet, target_sheet, headers_row, start_row):
    """
    Copia datos de una hoja de Excel a otra en el mismo archivo.
    
    Args:
        single_file (BytesIO): El archivo de Excel subido.
        source_sheet (str): El nombre de la hoja de origen.
        target_sheet (str): El nombre de la hoja de destino.
        headers_row (int): El número de fila donde están los encabezados en la hoja de destino.
        start_row (int): El número de fila donde se debe empezar a pegar.
        
    Returns:
        tuple: Un buffer de BytesIO con el archivo modificado y la cantidad de filas pegadas.
               Retorna (None, 0) si ocurre un error.
    """
    try:
        df_base = pd.read_excel(single_file, sheet_name=source_sheet, engine="openpyxl")
        
        wb = load_workbook(single_file)
        ws = wb[target_sheet]

        headers_plantilla = {
            str(cell.value).strip(): cell.column
            for cell in ws[int(headers_row)] if cell.value
        }
        
        columnas_comunes = [col for col in df_base.columns if col in headers_plantilla]
        
        if not columnas_comunes:
            raise ValueError("No se encontraron columnas coincidentes entre las hojas de origen y destino.")

        df_filtrado = df_base[columnas_comunes]

        rows_to_paste = dataframe_to_rows(df_filtrado, index=False, header=False)
        
        for r_idx, row in enumerate(rows_to_paste, start=int(start_row)):
            for col_name, value in zip(df_filtrado.columns, row):
                col_idx = headers_plantilla[col_name]
                ws.cell(row=r_idx, column=col_idx, value=value)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output, len(df_filtrado)

    except KeyError as e:
        st.error(f"Error: La hoja '{e.args[0]}' no existe en el archivo. Por favor, verifica el nombre.")
        return None,
